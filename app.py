from flask import Flask, render_template, request, redirect, url_for, session, send_file
import pandas as pd
import os

app = Flask(__name__)
app.secret_key = 'care_web_secret_key_123'  # セッション用の秘密鍵（適宜変更してください）

# ログイン用のユーザー情報（IDはadmin Passはpassword123）
USER_DATA = {"username": "admin", "password": "password123"}

DATA_FILE = "data/users.csv"
if not os.path.exists("data"):
    os.makedirs("data")

if not os.path.exists(DATA_FILE):
    df_init = pd.DataFrame(columns=["name", "care_manager", "care_level"])
    df_init.to_csv(DATA_FILE, index=False, encoding="utf-8")

SUPPORT_LEVELS = ["要支援１", "要支援２", "事業"]
CARE_LEVELS = ["要介護１", "要介護２", "要介護３", "要介護４", "要介護５"]

# --- ログインチェック用関数 ---
def is_logged_in():
    return session.get('logged_in')

# --- ルート定義 ---

@app.route('/')
def index():
    if not is_logged_in():
        return redirect(url_for('login'))
    return render_template("form.html")

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form['username'] == USER_DATA['username'] and \
           request.form['password'] == USER_DATA['password']:
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            return "ログイン失敗。やり直してください。<a href='/login'>戻る</a>"
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route("/dashboard")
def dashboard():
    if not is_logged_in():
        return redirect(url_for('login'))
    
    if not os.path.exists(DATA_FILE):
        return "データファイルがありません。"

    df = pd.read_csv(DATA_FILE)
    managers = sorted(df["care_manager"].unique().tolist())
    all_levels = SUPPORT_LEVELS + CARE_LEVELS

    matrix = {l: {m: [] for m in managers} for l in all_levels}
    counts = {l: {m: 0 for m in managers} for l in all_levels}
    
    for idx, row in df.iterrows():
        l, m = row["care_level"], row["care_manager"]
        if l in matrix and m in managers:
            matrix[l][m].append({"id": idx, "name": row["name"]})
            counts[l][m] += 1

    subtotal_support = {m: sum(counts[l][m] for l in SUPPORT_LEVELS) for m in managers}
    subtotal_care = {m: sum(counts[l][m] for l in CARE_LEVELS) for m in managers}
    grand_totals = {m: subtotal_support[m] + subtotal_care[m] for m in managers}

    row_totals = {l: sum(counts[l].values()) for l in all_levels}
    row_subtotal_support = sum(subtotal_support.values())
    row_subtotal_care = sum(subtotal_care.values())
    row_grand_total = sum(grand_totals.values())

    return render_template(
        "dashboard.html",
        managers=managers,
        matrix=matrix,
        counts=counts,
        support_levels=SUPPORT_LEVELS,
        care_levels=CARE_LEVELS,
        subtotal_support=subtotal_support,
        subtotal_care=subtotal_care,
        grand_totals=grand_totals,
        row_totals=row_totals,
        row_subtotal_support=row_subtotal_support,
        row_subtotal_care=row_subtotal_care,
        row_grand_total=row_grand_total
    )

@app.route("/submit", methods=["POST"])
def submit():
    if not is_logged_in(): return redirect(url_for('login'))
    df = pd.read_csv(DATA_FILE)
    new_row = {
        "name": request.form.get("name"),
        "care_manager": request.form.get("care_manager"),
        "care_level": request.form.get("care_level"),
    }
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df.to_csv(DATA_FILE, index=False, encoding="utf-8")
    return redirect(url_for('dashboard'))

@app.route("/delete/<int:user_id>")
def delete_user(user_id):
    if not is_logged_in(): return redirect(url_for('login'))
    df = pd.read_csv(DATA_FILE)
    df = df.drop(df.index[user_id]).reset_index(drop=True)
    df.to_csv(DATA_FILE, index=False, encoding="utf-8")
    return redirect(url_for('dashboard'))

@app.route("/edit/<int:user_id>")
def edit_page(user_id):
    if not is_logged_in(): return redirect(url_for('login'))
    df = pd.read_csv(DATA_FILE)
    user_data = df.iloc[user_id].to_dict()
    return render_template("edit.html", user=user_data, user_id=user_id)

@app.route("/update/<int:user_id>", methods=["POST"])
def update_user(user_id):
    if not is_logged_in(): return redirect(url_for('login'))
    df = pd.read_csv(DATA_FILE)
    df.at[user_id, "name"] = request.form.get("name")
    df.at[user_id, "care_manager"] = request.form.get("care_manager")
    df.at[user_id, "care_level"] = request.form.get("care_level")
    df.to_csv(DATA_FILE, index=False, encoding="utf-8")
    return redirect(url_for('dashboard'))

@app.route("/rename_manager", methods=["POST"])
def rename_manager():
    if not is_logged_in(): return redirect(url_for('login'))
    old_name = request.form.get("old_name")
    new_name = request.form.get("new_name")
    if old_name and new_name:
        df = pd.read_csv(DATA_FILE)
        df["care_manager"] = df["care_manager"].replace(old_name, new_name)
        df.to_csv(DATA_FILE, index=False, encoding="utf-8")
    return redirect(url_for('dashboard'))

# --- app.py の一番上の方に追加 ---
from openpyxl.styles import Alignment  # これが必要！

# --- app.py の download_excel 部分をこれに差し替え ---
@app.route("/download")
def download_excel():
    if not is_logged_in():
        return redirect(url_for('login'))
    
    if not os.path.exists(DATA_FILE):
        return "データがありません。"

    df = pd.read_csv(DATA_FILE)
    if df.empty:
        return "データが空です。"

    # 1. 介護度の定義
    support_levels = ["要支援１", "要支援２", "事業"]
    care_levels = ["要介護１", "要介護２", "要介護３", "要介護４", "要介護５"]
    all_levels = support_levels + care_levels

    # 2. 名前マトリックスと人数マトリックスの作成
    pivot_names = df.pivot_table(
        index='care_level', columns='care_manager', values='name', 
        aggfunc=lambda x: "\n".join(str(v) for v in x), fill_value=""
    ).reindex(all_levels).fillna("")

    pivot_counts = df.pivot_table(
        index='care_level', columns='care_manager', values='name', 
        aggfunc='count', fill_value=0
    ).reindex(all_levels).fillna(0)

    # 3. 横方向の「社内合計」を計算
    pivot_names['社内合計'] = pivot_counts.sum(axis=1).astype(int).map(lambda x: f"{x}")
    pivot_counts['社内合計'] = pivot_counts.sum(axis=1)

    # 4. Excel用データフレームの構築
    output_rows = []
    managers_with_total = pivot_names.columns.tolist() # 「社内合計」も含まれる

    for level in all_levels:
        # 名前行
        name_row = {m: pivot_names.loc[level, m] for m in managers_with_total}
        name_row['区分'] = level
        output_rows.append(name_row)
        
        # 人数行
        count_row = {m: f"{int(pivot_counts.loc[level, m])}名" for m in managers_with_total}
        count_row['区分'] = f"{level} 人数"
        output_rows.append(count_row)

    # 5. 縦方向の「総合計」行を作成して追加
    grand_total_row = {'区分': '総合計'}
    for m in managers_with_total:
        total_val = pivot_counts[m].sum()
        grand_total_row[m] = f"{int(total_val)}名"
    output_rows.append(grand_total_row)

    final_df = pd.DataFrame(output_rows)
    final_df = final_df[['区分'] + managers_with_total]

    # 6. Excel書き出し & 装飾
    excel_file = "居宅支援名簿_完全集計版.xlsx"
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        final_df.to_excel(writer, index=False, sheet_name='名簿')
        ws = writer.sheets['名簿']
        
        from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
        
        # 色の設定
        header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        count_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # 総合計は赤系
        
        # 罫線の設定
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                cell.border = border
                
                # 1行目（ヘッダー）
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = Font(bold=True)
                
                # A列（区分）
                label = str(ws.cell(row=cell.row, column=1).value)
                
                # 人数行に色
                if "人数" in label:
                    cell.fill = count_fill
                
                # 総合計行に色
                if "総合計" in label:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
                
                # 社内合計列（一番右）に太字
                if cell.column == ws.max_column:
                    cell.font = Font(bold=True)

    return send_file(excel_file, as_attachment=True)

if __name__ == "__main__":
    # Renderなどの環境ではPORT環境変数を使うためのおまじない
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)